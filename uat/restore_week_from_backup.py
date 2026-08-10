#!/usr/bin/env python3
"""Sinh SQL phục hồi một tuần đã xóa từ file backup Excel.

    python uat/restore_week_from_backup.py <file_backup.xlsx> [> restore.sql]

Công cụ này CỐ Ý không tự kết nối và không tự chạy. Nó chỉ in ra SQL để người
có quyền đọc lại, kiểm tra, rồi tự chạy trong một transaction. Phục hồi dữ liệu
là thao tác nhạy cảm — phải có người xem trước khi ghi.

SQL sinh ra:
  - chạy trong BEGIN/COMMIT, dừng ngay khi có lỗi;
  - dùng INSERT ... ON CONFLICT (id) DO UPDATE nên chạy lại nhiều lần vẫn cho
    cùng kết quả, không tạo bản trùng;
  - đặt lại deleted_at = null để đưa việc trở lại trạng thái hoạt động;
  - KHÔNG tự mở lại work_weeks. Sau khi phục hồi việc, người vận hành tự quyết
    định có chuyển tuần từ ARCHIVED về ACTIVE hay không.

Yêu cầu: openpyxl  (pip install openpyxl)
"""
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Cần openpyxl: pip install openpyxl")

# Nhãn cột trong file backup -> cột của public.jobs. Chỉ các cột NHẬP mới được
# phục hồi; các cột dẫn xuất (sản lượng, hòa vốn, chênh lệch...) do view tính lại.
COLUMNS = {
    "ID": "id", "Team ID": "team_id", "Ngày bắt đầu": "start_date", "Ngày kết thúc": "end_date",
    "Hạng mục Cấp 1": "category_name", "Nội dung công việc": "content", "Vị trí": "location",
    "Khối lượng": "quantity", "Tổ trưởng": "count_leader", "Thợ bậc 1": "count_worker1",
    "Thợ bậc 2": "count_worker2", "Thợ bậc 3": "count_worker3", "Thợ phụ": "count_helper",
    "Việc đặc biệt": "is_special_labor", "Mã nhóm": "group_code", "Week ID": "week_id",
    "Request key": "request_key", "Dòng nguồn cũ": "legacy_source_row",
    "Tạo lúc": "created_at", "Người tạo": "created_by",
}
NUMERIC = {"quantity", "count_leader", "count_worker1", "count_worker2", "count_worker3",
           "count_helper", "legacy_source_row"}
BOOLEAN = {"is_special_labor"}
NULLABLE_UUID = {"group_code", "week_id", "request_key", "created_by"}


def literal(column, value):
    if value is None or str(value).strip() == "":
        return "null"
    text = str(value).strip()
    if column in BOOLEAN:
        return "true" if text.lower() in ("true", "1", "có", "x") else "false"
    if column in NUMERIC:
        return text
    return "'" + text.replace("'", "''") + "'"


def main(path):
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = book["BACKUP_TUAN"] if "BACKUP_TUAN" in book.sheetnames else book.active
    rows = [r for r in sheet.iter_rows(values_only=True)
            if any(c is not None and str(c).strip() for c in r)]
    if len(rows) < 2:
        sys.exit("File backup không có dòng dữ liệu nào.")

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    missing = [label for label in COLUMNS if label not in header]
    if missing:
        sys.exit("File backup thiếu cột bắt buộc: " + ", ".join(missing)
                 + "\nFile này được tạo bởi phiên bản cũ và KHÔNG phục hồi đầy đủ được.")
    index = {label: header.index(label) for label in COLUMNS}
    targets = list(COLUMNS.values())

    print(f"-- Phục hồi từ {Path(path).name} — {len(rows) - 1} công việc.")
    print("-- Đọc kỹ trước khi chạy. Chạy lại nhiều lần cho cùng kết quả.")
    print("\\set ON_ERROR_STOP on")
    print("begin;")
    for line in rows[1:]:
        values = [literal(COLUMNS[label], line[index[label]]) for label in COLUMNS]
        assignments = ",\n    ".join(
            f"{c} = excluded.{c}" for c in targets if c != "id")
        print(f"""
insert into public.jobs ({", ".join(targets)}, deleted_at, updated_at, updated_by)
values ({", ".join(values)}, null, now(), auth.uid())
on conflict (id) do update set
    {assignments},
    deleted_at = null,
    updated_at = now();""")
    print("\n-- Kiểm tra trước khi commit:")
    print("select count(*) as viec_da_phuc_hoi from public.jobs")
    print(f"  where deleted_at is null and id in ({', '.join(literal('id', l[index['ID']]) for l in rows[1:])});")
    print("commit;")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        sys.exit(__doc__)
    main(sys.argv[1])
